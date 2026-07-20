#!/usr/bin/env python3
"""Optional local smoke check against gitignored 1.xlsx/2.xlsx/3.xlsx.

Prints ONLY aggregate counts. Never prints material codes or SO samples.
Requires openpyxl if real xlsx files are present; otherwise exits 0 with skip.
"""

from __future__ import annotations

import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[3]
FILES = [ROOT / "1.xlsx", ROOT / "2.xlsx", ROOT / "3.xlsx"]

COL_E1_SO = 0
COL_E2_SO = 1
COL_E2_MOTHER = 5
COL_E3_MOTHER = 4
COL_E3_CHILD = 8


def _cell(row: tuple[object, ...], index: int) -> str:
    if index >= len(row):
        return ""
    value = row[index]
    return "" if value is None else str(value).strip()


def main() -> int:
    missing = [str(path.name) for path in FILES if not path.is_file()]
    if missing:
        print(f"SKIP real-file smoke: missing {', '.join(missing)}")
        return 0

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("SKIP real-file smoke: openpyxl not installed")
        return 0

    counts: dict[str, int] = {}
    for path in FILES:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        row_count = 0
        for i, _row in enumerate(ws.iter_rows(values_only=True)):
            if i == 0:
                continue
            row_count += 1
        counts[path.name] = row_count
        wb.close()

    wb1 = load_workbook(FILES[0], read_only=True, data_only=True)
    ws1 = wb1.active
    so_keys: set[str] = set()
    for i, row in enumerate(ws1.iter_rows(values_only=True)):
        if i == 0:
            continue
        so = _cell(row, COL_E1_SO)
        if so:
            so_keys.add(so)
    wb1.close()

    wb2 = load_workbook(FILES[1], read_only=True, data_only=True)
    ws2 = wb2.active
    demand_keys = 0
    mothers_in_demand: set[str] = set()
    for i, row in enumerate(ws2.iter_rows(values_only=True)):
        if i == 0:
            continue
        so = _cell(row, COL_E2_SO)
        mother = _cell(row, COL_E2_MOTHER)
        if so and mother:
            demand_keys += 1
            mothers_in_demand.add(mother)
    wb2.close()

    wb3 = load_workbook(FILES[2], read_only=True, data_only=True)
    ws3 = wb3.active
    mothers_in_bom: set[str] = set()
    child_links = 0
    for i, row in enumerate(ws3.iter_rows(values_only=True)):
        if i == 0:
            continue
        mother = _cell(row, COL_E3_MOTHER)
        child = _cell(row, COL_E3_CHILD)
        if mother:
            mothers_in_bom.add(mother)
        if mother and child:
            child_links += 1
    wb3.close()

    mother_overlap = len(mothers_in_demand & mothers_in_bom)

    print("=== Real-file smoke (aggregates only) ===")
    print(f"excel1_data_rows={counts['1.xlsx']}")
    print(f"excel2_data_rows={counts['2.xlsx']}")
    print(f"excel3_data_rows={counts['3.xlsx']}")
    print(f"excel1_distinct_so_count={len(so_keys)}")
    print(f"excel2_rows_with_so_and_mother={demand_keys}")
    print(f"excel2_distinct_mother_count={len(mothers_in_demand)}")
    print(f"excel3_distinct_mother_count={len(mothers_in_bom)}")
    print(f"excel3_mother_child_link_rows={child_links}")
    print(f"mother_overlap_excel2_and_excel3={mother_overlap}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
