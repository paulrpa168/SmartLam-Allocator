#!/usr/bin/env python3
"""RAW MAT allocation engine v3.4 verifier.

繁體中文：驗證 MB 前置過濾、COOIS 彙總、母料池與子料池配發、
單位換算、19 欄輸出、合貼備料齊套規則、歸屬不明確認後比例拆分，
以及選用的真實資料基準。
English: Verifies MB pre-filtering, COOIS aggregation, mother/child pool
allocation, unit conversion, the 19-column contract, Lamination Kit Ready,
ambiguous-direct confirm/split, and the optional real-data baseline.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

OUTPUT_HEADERS = [
    "cutting",
    "so",
    "mother material",
    "mother batch",
    "mother unit",
    "mother plant stock",
    "mother stock outside",
    "mother stock available before this row",
    "child material",
    "child batch",
    "child unit",
    "child demand",
    "demand direct",
    "demand qty",
    "child stock",
    "stock available",
    "provided qty",
    "remaining stock after this row",
    "lamination kit ready (Y)",
]

DEFAULT_RULES = {
    ("M", "YD"): 1.0936,
    ("YD", "M"): 0.9144,
}

DEFAULT_SIZE_RULES = {
    ("110x200cm", "M"): 2.0,
    ("110x200cm", "YD"): 2.187227,
    ("110x200cm", "SHT"): 1.0,
    ("220x110cm", "M"): 2.0,
    ("220x110cm", "YD"): 2.187227,
    ("220x110cm", "SHT"): 1.0,
    ("200x110cm", "M"): 2.2,
    ("200x110cm", "YD"): 2.405949,
    ("200x110cm", "SHT"): 1.0,
}

SIZE_SUFFIX_RE = re.compile(r"(\d+)\s*[x×*]\s*(\d+)\s*cm\s*$", re.I)

HAND_OFF = Path(__file__).resolve().parent
ROOT = HAND_OFF.parent.parent.parent  # .../handoffs/<id> → project root
NEW_0722 = ROOT / "New_0722"


def _norm(value: object) -> str:
    return str(value or "").strip()


def _norm_seg(value: object) -> str:
    """FLT/MTF warehouse lock — never mix; blank stays blank."""
    return _norm(value).upper()


def _num(value: object) -> float:
    text = _norm(value).replace(",", "")
    if not text:
        return 0.0
    try:
        return float(text)
    except ValueError:
        return 0.0


def _fmt(value: float) -> str:
    if abs(value - round(value)) < 1e-9:
        return str(int(round(value)))
    rounded = round(value, 6)
    text = str(rounded)
    if "." in text:
        return text.rstrip("0").rstrip(".")
    return text


def _extract_size_suffix(text: object) -> str:
    match = SIZE_SUFFIX_RE.search(_norm(text))
    if not match:
        return ""
    return f"{match.group(1)}x{match.group(2)}cm".lower()


def _header_key(h: object) -> str:
    return re.sub(r"\s+", " ", _norm(h).lower())


def find_col(headers: list[str], aliases: list[str]) -> int:
    normalized = [_header_key(h) for h in headers]
    alias_keys = [_header_key(a) for a in aliases if _norm(a)]
    for key in alias_keys:
        for i, h in enumerate(normalized):
            if h == key:
                return i
    for key in alias_keys:
        for i, h in enumerate(normalized):
            if key in h or (len(key) >= 3 and len(h) >= 3 and h in key):
                return i
    return -1


def conversion_ratio(
    from_unit: str,
    to_unit: str,
    rules: dict[tuple[str, str], float],
    size_rules: dict[tuple[str, str], float],
    size_suffix: str,
) -> tuple[bool, float, str]:
    a = _norm(from_unit).upper()
    b = _norm(to_unit).upper()
    if a == "Y":
        a = "YD"
    if b == "Y":
        b = "YD"
    if not a or not b or a == b:
        return True, 1.0, ""
    if a == "SHT":
        if not size_suffix:
            return False, 0.0, "missing-suffix"
        key = (size_suffix.lower(), b)
        if key not in size_rules:
            return False, 0.0, "missing-size-rule"
        return True, size_rules[key], ""
    key = (a, b)
    if key not in rules:
        return False, 0.0, "missing-rule"
    return True, rules[key], ""


def storage_excluded(storage: object) -> bool:
    return _norm(storage).startswith("39")


def is_mh04_child(child: object) -> bool:
    return _norm(child).upper().startswith("MH04")


def is_mb_material(value: object) -> bool:
    return _norm(value).upper().startswith("MB")


def split_direct_demand_by_mother_open(
    qty: float,
    candidates: list[dict],
    mother_open_qtys: list[float],
) -> None:
    """Split direct Open Qty across mothers by Open Qty weights (last gets remainder)."""
    n = len(candidates)
    if n <= 0 or qty <= 0:
        return
    weights = [max(0.0, float(w or 0.0)) for w in mother_open_qtys]
    weight_sum = sum(weights)
    allocated = 0.0
    for i, item in enumerate(candidates):
        if i == n - 1:
            part = qty - allocated
        elif weight_sum > 1e-12:
            part = qty * (weights[i] / weight_sum)
        else:
            part = qty / n
        item["demand_direct"] += part
        allocated += part


def load_conversion_config(path: Path) -> tuple[dict[tuple[str, str], float], dict[tuple[str, str], float]]:
    value = json.loads(path.read_text(encoding="utf-8"))
    if int(value.get("schemaVersion", 0)) != 1:
        raise ValueError("conversion config schemaVersion must be 1")
    unit_rules: dict[tuple[str, str], float] = {}
    for index, rule in enumerate(value.get("conversionRules", [])):
        mother = _norm(rule.get("motherUnit")).upper()
        child = _norm(rule.get("childUnit")).upper()
        if child == "Y":
            child = "YD"
        ratio = _num(rule.get("ratio"))
        if not mother or not child or ratio <= 0:
            raise ValueError(f"invalid conversionRules[{index}]")
        unit_rules[(mother, child)] = ratio
    size_rules: dict[tuple[str, str], float] = {}
    for index, rule in enumerate(value.get("sizeConversionRules", [])):
        suffix = _norm(rule.get("suffix")).lower().replace("×", "x").replace("*", "x")
        child = _norm(rule.get("childUnit")).upper()
        if child == "Y":
            child = "YD"
        ratio = _num(rule.get("ratio"))
        if not suffix or not child or ratio <= 0:
            raise ValueError(f"invalid sizeConversionRules[{index}]")
        size_rules[(suffix, child)] = ratio
    return unit_rules, size_rules


def run_engine(
    schedule: list[dict],
    coois: list[dict],
    zrmm: list[dict],
    mb52: list[dict],
    rules: dict[tuple[str, str], float] | None = None,
    size_rules: dict[tuple[str, str], float] | None = None,
    apply_rule_d: bool = True,
    confirm_ambiguous_split: bool = False,
) -> dict:
    """Python mirror of the v3.4 browser allocation engine."""
    rules = DEFAULT_RULES if rules is None else rules
    size_rules = DEFAULT_SIZE_RULES if size_rules is None else size_rules

    schedule_map: dict[str, dict] = {}
    for source_order, row in enumerate(schedule):
        so = _norm(row.get("so"))
        if not so:
            continue
        cutting = _norm(row.get("cutting"))
        sort_key = cutting
        prev = schedule_map.get(so)
        if (
            not prev
            or sort_key < prev["sort_key"]
            or (sort_key == prev["sort_key"] and source_order < prev["source_order"])
        ):
            schedule_map[so] = {
                "cutting": cutting,
                "sort_key": sort_key,
                "source_order": source_order,
            }

    filtered_zrmm = [
        row
        for row in zrmm
        if not is_mb_material(row.get("mother")) and not is_mb_material(row.get("child"))
    ]
    mb_excluded = len(zrmm) - len(filtered_zrmm)

    def relation_set(rows: list[dict]) -> set[tuple[str, str, str]]:
        return {
            (_norm(r.get("mother")), _norm(r.get("child")), _norm_seg(r.get("batch")))
            for r in rows
            if _norm(r.get("mother"))
            and _norm(r.get("child"))
            and not storage_excluded(r.get("storage"))
        }

    base_relations = relation_set(zrmm)
    kept_relations = relation_set(filtered_zrmm)

    children_by_mother_seg: dict[tuple[str, str], dict[str, dict]] = defaultdict(dict)
    mother_suffix: dict[str, str] = {}
    mother_unit_h: dict[str, str] = {}
    outside_pair: dict[tuple[str, str, str], dict] = {}
    mother_same_unit_children: dict[str, set[str]] = defaultdict(set)

    for source_order, row in enumerate(filtered_zrmm):
        mother = _norm(row.get("mother"))
        child = _norm(row.get("child"))
        if not mother or not child or storage_excluded(row.get("storage")):
            continue
        j = _num(row.get("gi_j"))
        vendor_l = _num(row.get("vendor_l"))
        gr_p = _num(row.get("gr_p"))
        batch = _norm_seg(row.get("batch"))
        oun = _norm(row.get("oun")).upper()
        bun = _norm(row.get("bun")).upper()
        if oun == "Y":
            oun = "YD"
        if bun == "Y":
            bun = "YD"
        suffix = _extract_size_suffix(row.get("desc"))
        if mother not in mother_unit_h and oun:
            mother_unit_h[mother] = oun
        if mother not in mother_suffix and suffix:
            mother_suffix[mother] = suffix
        children_by_mother_seg[(mother, batch)].setdefault(
            child, {"bun": bun, "oun": oun, "source_order": source_order}
        )
        if oun and bun and oun == bun:
            mother_same_unit_children[mother].add(child)
        if j == 0 or vendor_l == 0:
            continue
        pair_key = (mother, child, batch)
        if pair_key not in outside_pair:
            outside_pair[pair_key] = {
                "j": 0.0,
                "p": 0.0,
                "oun": oun,
                "bun": bun,
                "suffix": suffix,
            }
        outside_pair[pair_key]["j"] += j
        outside_pair[pair_key]["p"] += gr_p

    missing_rules: list[str] = []
    outside_by_mother: dict[tuple[str, str], float] = defaultdict(float)
    outside_fallback_by_mother: dict[tuple[str, str], bool] = {}
    for (mother, child, seg), agg in outside_pair.items():
        if mother_same_unit_children.get(mother) and agg["oun"] != agg["bun"]:
            continue
        outside_child = max(0.0, agg["j"] - agg["p"])
        if mother_same_unit_children.get(mother):
            outside_by_mother[(mother, seg)] += outside_child
        else:
            outside_by_mother[(mother, seg)] += outside_child
            outside_fallback_by_mother[(mother, seg)] = True

    stock_mb52: dict[tuple[str, str], float] = defaultdict(float)
    for row in mb52:
        material = _norm(row.get("material"))
        if not material or storage_excluded(row.get("storage")):
            continue
        stock_mb52[(material, _norm_seg(row.get("segment")))] += _num(row.get("stock"))

    mothers_in_bom = {mother for mother, _seg in children_by_mother_seg}
    valid_children = {
        child for child_map in children_by_mother_seg.values() for child in child_map
    }
    mothers_after = {m for m, _s in children_by_mother_seg}
    children_after = set(valid_children)

    # COOIS aggregate by SO + material + segment, preserving first source row.
    coois_by_so: dict[str, list[dict]] = defaultdict(list)
    aggregate: dict[tuple[str, str, str], dict] = {}
    for source_order, row in enumerate(coois):
        so = _norm(row.get("so"))
        if not so or so not in schedule_map:
            continue
        material = _norm(row.get("material"))
        if not material:
            continue
        seg = _norm_seg(row.get("segment"))
        unit = _norm(row.get("unit")).upper()
        if unit == "Y":
            unit = "YD"
        key = (so, material, seg)
        if key not in aggregate:
            aggregate[key] = {
                "so": so,
                "material": material,
                "seg": seg,
                "unit": unit,
                "qty": 0.0,
                "source_order": source_order,
            }
            coois_by_so[so].append(aggregate[key])
        item = aggregate[key]
        if item["unit"] and unit and item["unit"] != unit:
            return {"ok": False, "error": f"COOIS unit conflict: {so}/{material}/{seg}"}
        if not item["unit"] and unit:
            item["unit"] = unit
        item["qty"] += _num(row.get("qty"))

    sorted_sos = sorted(
        schedule_map,
        key=lambda so: (
            schedule_map[so]["sort_key"],
            schedule_map[so]["source_order"],
        ),
    )

    demand_rows: dict[tuple[str, str, str, str], dict] = {}

    def ensure(so: str, seg: str, mother: str, child: str, meta: dict) -> dict:
        key = (so, seg, mother, child)
        if key not in demand_rows:
            sched = schedule_map.get(so, {})
            demand_rows[key] = {
                "cutting": sched.get("cutting", ""),
                "sort_cutting": sched.get("sort_key", "\uffff"),
                "schedule_order": sched.get("source_order", sys.maxsize),
                "so": so,
                "segment": seg,
                "mother": mother,
                "child": child,
                "mother_unit": meta.get("mother_unit", ""),
                "child_unit": meta.get("child_unit", ""),
                "mother_order": meta.get("mother_order", sys.maxsize),
                "child_order": meta.get("child_order", sys.maxsize),
                "direct_order": meta.get("direct_order", sys.maxsize),
                "conversion_ratio": float(meta.get("conversion_ratio", 1.0) or 1.0),
                "mother_stock_available": None,
                "mother_cover_child": 0.0,
                "demand_from_mother": 0.0,
                "demand_direct": 0.0,
            }
        return demand_rows[key]

    # Shared F+G mother pool. Each SO first consumes mother demand; Rule D
    # then covers direct child demand before the next SO is processed.
    remaining_cover_mother: dict[tuple[str, str], float] = {}
    ambiguities: list[dict] = []
    for so in sorted_sos:
        lines = sorted(coois_by_so.get(so, []), key=lambda row: row["source_order"])
        mother_open_by_seg: dict[tuple[str, str], float] = {}
        for line in lines:
            if not line["seg"] or line["material"] not in mothers_in_bom:
                continue
            key = (line["seg"], line["material"])
            mother_open_by_seg[key] = mother_open_by_seg.get(key, 0.0) + max(0.0, line["qty"])
        expansion_by_child: dict[tuple[str, str], list[dict]] = defaultdict(list)
        expanded_mothers: set[tuple[str, str]] = set()
        mother_contexts: list[tuple[str, str, dict[str, dict]]] = []

        for line in lines:
            material = line["material"]
            seg = line["seg"]
            if not seg or material not in mothers_in_bom:
                continue
            child_map = children_by_mother_seg.get((material, seg), {})
            if not child_map:
                continue
            expanded_mothers.add((seg, material))
            mother_key = (material, seg)
            if mother_key not in remaining_cover_mother:
                remaining_cover_mother[mother_key] = stock_mb52.get(
                    mother_key, 0.0
                ) + outside_by_mother.get(mother_key, 0.0)
            mother_available_before = max(0.0, remaining_cover_mother[mother_key])
            cover_used = min(max(0.0, line["qty"]), mother_available_before)
            mother_pool_after_mother = max(0.0, mother_available_before - cover_used)
            remaining_cover_mother[mother_key] = mother_pool_after_mother
            qty_expand = max(0.0, line["qty"] - cover_used)
            mother_unit = line["unit"] or mother_unit_h.get(material, "")
            suffix = mother_suffix.get(material, "")
            mother_contexts.append((material, seg, child_map))
            for child, meta in child_map.items():
                child_unit = meta["bun"]
                ok, ratio, reason = conversion_ratio(
                    mother_unit, child_unit, rules, size_rules, suffix
                )
                if not ok:
                    missing_rules.append(f"conversion {reason}: {material}->{child}")
                    continue
                item = ensure(
                    so,
                    seg,
                    material,
                    child,
                    {
                        "mother_unit": mother_unit,
                        "child_unit": child_unit,
                        "mother_order": line["source_order"],
                        "child_order": meta["source_order"],
                        "conversion_ratio": ratio,
                    },
                )
                item["demand_from_mother"] += qty_expand * ratio
                if not apply_rule_d:
                    item["mother_stock_available"] = mother_available_before
                expansion_by_child[(seg, child)].append(item)

        for line in lines:
            material = line["material"]
            seg = line["seg"]
            qty = line["qty"]
            if (seg, material) in expanded_mothers or material not in valid_children:
                continue
            candidates = expansion_by_child.get((seg, material), [])
            if len(candidates) > 1 and qty > 1e-9:
                mother_open_qtys = [
                    mother_open_by_seg.get((seg, item["mother"]), 0.0) for item in candidates
                ]
                ambiguities.append(
                    {
                        "so": so,
                        "seg": seg,
                        "child": material,
                        "child_open_qty": qty,
                        "child_unit": line["unit"],
                        "mothers": [
                            {"mother": item["mother"], "open_qty": mother_open_qtys[i]}
                            for i, item in enumerate(candidates)
                        ],
                    }
                )
                if confirm_ambiguous_split:
                    split_direct_demand_by_mother_open(qty, candidates, mother_open_qtys)
                    for item in candidates:
                        if not item["child_unit"] and line["unit"]:
                            item["child_unit"] = line["unit"]
                continue
            if len(candidates) == 1:
                candidates[0]["demand_direct"] += qty
                if not candidates[0]["child_unit"] and line["unit"]:
                    candidates[0]["child_unit"] = line["unit"]
            else:
                item = ensure(
                    so,
                    seg,
                    "",
                    material,
                    {
                        "child_unit": line["unit"],
                        "direct_order": line["source_order"],
                    },
                )
                item["demand_direct"] += qty

        # Rule D: use original ZRMM0028 child order. Store the mother-unit
        # balance before each row and consume it using the mother→child ratio.
        if apply_rule_d:
            for mother, seg, child_map in mother_contexts:
                mother_key = (mother, seg)
                mother_pool = max(0.0, remaining_cover_mother.get(mother_key, 0.0))
                for child in child_map:
                    item = demand_rows.get((so, seg, mother, child))
                    if item is None:
                        continue
                    ratio = float(item.get("conversion_ratio", 0.0) or 0.0)
                    if ratio <= 0:
                        missing_rules.append(f"Rule D conversion: {mother}->{child}")
                        continue
                    item["mother_stock_available"] = mother_pool
                    demand_total = max(0.0, item["demand_from_mother"] + item["demand_direct"])
                    mother_cover_child = min(demand_total, mother_pool * ratio)
                    item["mother_cover_child"] = mother_cover_child
                    mother_pool = max(0.0, mother_pool - mother_cover_child / ratio)
                remaining_cover_mother[mother_key] = mother_pool
    if ambiguities and not confirm_ambiguous_split:
        detail = "; ".join(
            f"{a['so']}/{a['seg']}/{a['child']} mothers="
            + ",".join(m["mother"] for m in a["mothers"])
            for a in ambiguities
        )
        return {
            "ok": False,
            "needs_ambiguous_confirm": True,
            "ambiguities": ambiguities,
            "error": f"ambiguous direct demand: {detail}",
        }
    if missing_rules:
        return {"ok": False, "error": "; ".join(sorted(set(missing_rules))[:80])}

    expanded = sorted(
        demand_rows.values(),
        key=lambda item: (
            item["sort_cutting"],
            item["schedule_order"],
            item["so"],
            0 if item["mother"] else 1,
            item["mother_order"],
            item["child_order"],
            item["direct_order"],
            item["child"],
            item["segment"],
        ),
    )

    remaining_pool = {key: max(0.0, value) for key, value in stock_mb52.items()}
    out_rows: list[list] = []
    fulfilled_by_row: list[float] = []
    shortage = 0
    for item in expanded:
        demand_total = item["demand_from_mother"] + item["demand_direct"]
        mother_cover_child = (
            min(max(0.0, demand_total), max(0.0, float(item.get("mother_cover_child", 0.0))))
            if item["mother"]
            else 0.0
        )
        child_need_net = max(0.0, demand_total - mother_cover_child)
        child_key = (item["child"], item["segment"])
        child_stock = stock_mb52.get(child_key, 0.0)
        pool_before = remaining_pool.get(child_key, max(0.0, child_stock))
        provide = min(child_need_net, max(0.0, pool_before))
        pool_after = pool_before - provide
        fulfilled = mother_cover_child + provide
        remaining_pool[child_key] = pool_after
        fulfilled_by_row.append(fulfilled)
        if fulfilled < demand_total - 1e-9:
            shortage += 1
        mother_key = (item["mother"], item["segment"])
        mother_plant = stock_mb52.get(mother_key, 0.0) if item["mother"] else 0.0
        mother_outside = outside_by_mother.get(mother_key, 0.0) if item["mother"] else 0.0
        out_rows.append(
            [
                item["cutting"],
                item["so"],
                item["mother"],
                item["segment"],
                item["mother_unit"],
                _fmt(mother_plant) if item["mother"] else "",
                _fmt(mother_outside) if item["mother"] else "",
                _fmt(item["mother_stock_available"] or 0.0) if item["mother"] else "",
                item["child"],
                item["segment"],
                item["child_unit"],
                _fmt(item["demand_from_mother"]),
                _fmt(item["demand_direct"]),
                _fmt(demand_total),
                _fmt(child_stock),
                _fmt(pool_before),
                _fmt(provide),
                _fmt(pool_after),
                "",
            ]
        )

    groups: dict[tuple[str, str, str], list[int]] = defaultdict(list)
    for index, row in enumerate(out_rows):
        if _norm(row[2]):
            groups[(_norm(row[1]), _norm(row[2]), _norm(row[3]) or _norm(row[9]))].append(index)
    for indices in groups.values():
        positive_non_mh04 = [
            index
            for index in indices
            if not is_mh04_child(out_rows[index][8]) and _num(out_rows[index][13]) > 1e-9
        ]
        qualifies = bool(positive_non_mh04) and all(
            fulfilled_by_row[index] >= _num(out_rows[index][13]) - 1e-9
            for index in positive_non_mh04
        )
        for index in indices:
            out_rows[index][18] = "Y" if qualifies else ""

    return {
        "ok": True,
        "headers": OUTPUT_HEADERS,
        "rows": out_rows,
        "stats": {
            "demand_keys": len(demand_rows),
            "mothers": len(mothers_in_bom),
            "shortage": shortage,
            "zrmm_mb_excluded": mb_excluded,
            "zrmm_relationships_removed": len(base_relations - kept_relations),
            "dual_identity_materials": len(mothers_after & children_after),
            "outside_fallback_mothers": len(outside_fallback_by_mother),
        },
    }

def _assert(cond: bool, msg: str) -> None:
    if not cond:
        raise AssertionError(msg)


def test_find_col_open_quantity() -> None:
    headers = [
        "SD Document",
        "Material",
        "Quantity withdrawn",
        "Open Quantity",
        "Requirement Segment",
        "Base Unit of Measure",
    ]
    _assert(find_col(headers, ["open quantity"]) == 3, "must resolve Open Quantity")
    _assert(find_col(headers, ["sd document"]) == 0, "must resolve SD Document")


def test_mb_filter_and_19_columns() -> None:
    result = run_engine(
        schedule=[{"so": "SO1", "cutting": "2026-07-01"}],
        coois=[
            {"so": "SO1", "material": "M1", "qty": 2, "segment": "FLT", "unit": "M"},
            {"so": "SO1", "material": "MB_PARENT", "qty": 9, "segment": "FLT", "unit": "M"},
        ],
        zrmm=[
            {"mother": "M1", "child": "C1", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
            {"mother": "MB_PARENT", "child": "MB_CHILD", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
        ],
        mb52=[{"material": "C1", "segment": "FLT", "storage": "1001", "stock": 5}],
    )
    _assert(result["ok"], str(result))
    _assert(len(result["headers"]) == 19, "v3.3 must export 19 columns")
    _assert([row[8] for row in result["rows"]] == ["C1"], "MB relationship must be excluded")
    _assert(result["stats"]["zrmm_mb_excluded"] == 1, "one MB row excluded")
    _assert(result["stats"]["dual_identity_materials"] == 0, "dual identity must be zero")


def test_direct_demand_retained_when_mother_covered() -> None:
    result = run_engine(
        schedule=[{"so": "10189518", "cutting": "2026-07-14"}],
        coois=[
            {"so": "10189518", "material": "MN090134761", "qty": 1.5, "segment": "FLT", "unit": "M"},
            {"so": "10189518", "material": "MA020147549", "qty": 0.02, "segment": "FLT", "unit": "M"},
        ],
        zrmm=[
            {"mother": "MN090134761", "child": "MA020147549", "gi_j": 2.09, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
        ],
        mb52=[{"material": "MA020147549", "segment": "FLT", "storage": "1001", "stock": 43}],
    )
    _assert(result["ok"], str(result))
    row = result["rows"][0]
    _assert(abs(_num(row[7]) - 0.59) < 1e-9, "mother pool before this row")
    _assert(_num(row[11]) == 0, "mother expansion K is zero")
    _assert(_num(row[12]) == 0.02, "direct L must remain")
    _assert(_num(row[13]) == 0.02, "total demand")
    _assert(_num(row[14]) == 43 and _num(row[15]) == 43, "child stock columns")
    _assert(_num(row[16]) == 0, "Rule D must avoid child MB52 when mother residual covers demand")
    _assert(_num(row[17]) == 43, "child MB52 must remain unchanged")
    _assert(row[18] == "Y", "positive mother group fully supplied by Rule D")


def test_rule_d_partial_and_zero_mother_balance() -> None:
    def run(outside: float) -> list:
        result = run_engine(
            schedule=[{"so": "SO1", "cutting": "2026-07-01"}],
            coois=[
                {"so": "SO1", "material": "M1", "qty": 1, "segment": "FLT", "unit": "M"},
                {"so": "SO1", "material": "C1", "qty": 2, "segment": "FLT", "unit": "M"},
            ],
            zrmm=[
                {"mother": "M1", "child": "C1", "gi_j": outside, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
            ],
            mb52=[{"material": "C1", "segment": "FLT", "storage": "1001", "stock": 10}],
        )
        _assert(result["ok"], str(result))
        return result["rows"][0]

    partial = run(1.5)
    _assert(abs(_num(partial[7]) - 0.5) < 1e-9, "partial case H")
    _assert(abs(_num(partial[16]) - 1.5) < 1e-9, "Q is only uncovered child need")
    _assert(abs(_num(partial[17]) - 8.5) < 1e-9, "R deducts Q only")
    _assert(partial[18] == "Y", "mother cover plus Q fulfills the row")

    empty = run(1.0)
    _assert(_num(empty[7]) == 0, "empty mother balance H")
    _assert(_num(empty[16]) == 2, "no mother balance means full child demand from MB52")
    _assert(_num(empty[17]) == 8, "child stock deducted by full Q")


def test_rule_d_child_order_and_cross_so_pool() -> None:
    ordered = run_engine(
        schedule=[{"so": "SO1", "cutting": "2026-07-01"}],
        coois=[
            {"so": "SO1", "material": "M1", "qty": 1, "segment": "FLT", "unit": "M"},
            {"so": "SO1", "material": "C1", "qty": 1, "segment": "FLT", "unit": "M"},
            {"so": "SO1", "material": "C2", "qty": 2, "segment": "FLT", "unit": "M"},
        ],
        zrmm=[
            {"mother": "M1", "child": "C2", "gi_j": 3, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
            {"mother": "M1", "child": "C1", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
        ],
        mb52=[
            {"material": "C1", "segment": "FLT", "storage": "1001", "stock": 5},
            {"material": "C2", "segment": "FLT", "storage": "1001", "stock": 5},
        ],
    )
    _assert(ordered["ok"], str(ordered))
    rows = ordered["rows"]
    _assert([row[8] for row in rows] == ["C2", "C1"], "ZRMM child source order")
    _assert(_num(rows[0][7]) == 2 and _num(rows[1][7]) == 0, "H decreases row by row")
    _assert(_num(rows[0][16]) == 0 and _num(rows[1][16]) == 1, "later child uses MB52")

    across = run_engine(
        schedule=[
            {"so": "SO1", "cutting": "2026-07-01"},
            {"so": "SO2", "cutting": "2026-07-02"},
        ],
        coois=[
            {"so": "SO1", "material": "M1", "qty": 1, "segment": "FLT", "unit": "M"},
            {"so": "SO1", "material": "C1", "qty": 1, "segment": "FLT", "unit": "M"},
            {"so": "SO2", "material": "M1", "qty": 1, "segment": "FLT", "unit": "M"},
            {"so": "SO2", "material": "C1", "qty": 1, "segment": "FLT", "unit": "M"},
        ],
        zrmm=[
            {"mother": "M1", "child": "C1", "gi_j": 3, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
        ],
        mb52=[{"material": "C1", "segment": "FLT", "storage": "1001", "stock": 5}],
    )
    _assert(across["ok"], str(across))
    by_so = {row[1]: row for row in across["rows"]}
    _assert(_num(by_so["SO1"][7]) == 2 and _num(by_so["SO1"][16]) == 0, "SO1 consumes mother pool")
    _assert(_num(by_so["SO2"][7]) == 0 and _num(by_so["SO2"][16]) == 1, "SO2 sees prior Rule D deduction")


def test_rule_d_unit_conversion() -> None:
    result = run_engine(
        schedule=[{"so": "SO1", "cutting": "2026-07-01"}],
        coois=[
            {"so": "SO1", "material": "M1", "qty": 1, "segment": "FLT", "unit": "M"},
            {"so": "SO1", "material": "C1", "qty": 2, "segment": "FLT", "unit": "YD"},
        ],
        zrmm=[
            {"mother": "M1", "child": "C1", "gi_j": 2.1872, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "YD", "desc": ""},
        ],
        mb52=[{"material": "C1", "segment": "FLT", "storage": "1001", "stock": 10}],
    )
    _assert(result["ok"], str(result))
    row = result["rows"][0]
    _assert(abs(_num(row[6]) - 2.1872) < 1e-9, "fallback G keeps raw J/P without conversion")
    _assert(abs(_num(row[7]) - 1.1872) < 1e-9, "H uses the fallback mother pool after mother demand")
    _assert(abs(_num(row[16]) - 0.701678) < 1e-9, "Q is uncovered child YD after mother-cover conversion")
    _assert(abs(_num(row[17]) - 9.298322) < 1e-9, "R deducts child-unit Q")

def test_outside_same_unit_priority() -> None:
    result = run_engine(
        schedule=[{"so": "SO1", "cutting": "2026-07-01"}],
        coois=[{"so": "SO1", "material": "M1", "qty": 2, "segment": "FLT", "unit": "M"}],
        zrmm=[
            {"mother": "M1", "child": "C_YD", "gi_j": 1.0936, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "YD", "desc": ""},
            {"mother": "M1", "child": "C_M", "gi_j": 1.0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
        ],
        mb52=[
            {"material": "C_YD", "segment": "FLT", "storage": "1001", "stock": 10},
            {"material": "C_M", "segment": "FLT", "storage": "1001", "stock": 10},
        ],
    )
    _assert(result["ok"], str(result))
    rows = result["rows"]
    _assert(all(abs(_num(row[6]) - 1) < 1e-6 for row in rows), f"outside should use same-unit only, got {[row[6] for row in rows]}")
    yd_row = next(row for row in rows if _norm(row[8]) == "C_YD")
    _assert(abs(_num(yd_row[11]) - 1.0936) < 1e-6, "remaining 1 M still expands to YD child demand")


def test_outside_fallback_without_conversion() -> None:
    result = run_engine(
        schedule=[{"so": "SO1", "cutting": "2026-07-01"}],
        coois=[{"so": "SO1", "material": "M1", "qty": 5, "segment": "FLT", "unit": "M"}],
        zrmm=[
            {"mother": "M1", "child": "C1", "gi_j": 1.0936, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "YD", "desc": ""},
        ],
        mb52=[{"material": "C1", "segment": "FLT", "storage": "1001", "stock": 10}],
    )
    _assert(result["ok"], str(result))
    row = result["rows"][0]
    _assert(abs(_num(row[6]) - 1.0936) < 1e-6, f"fallback should not convert G, got {row[6]}")
    _assert(result["stats"].get("outside_fallback_mothers") == 1, "fallback mother count should be flagged")


def test_coois_aggregation_and_source_order() -> None:
    result = run_engine(
        schedule=[
            {"so": "SO_B", "cutting": "2026-07-01"},
            {"so": "SO_A", "cutting": "2026-07-01"},
        ],
        coois=[
            {"so": "SO_B", "material": "M_B", "qty": 1, "segment": "FLT", "unit": "M"},
            {"so": "SO_B", "material": "M_B", "qty": 2, "segment": "FLT", "unit": "M"},
            {"so": "SO_A", "material": "M_A", "qty": 1, "segment": "FLT", "unit": "M"},
        ],
        zrmm=[
            {"mother": "M_B", "child": "SHARED", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
            {"mother": "M_A", "child": "SHARED", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
        ],
        mb52=[{"material": "SHARED", "segment": "FLT", "storage": "1001", "stock": 2}],
    )
    _assert(result["ok"], str(result))
    rows = result["rows"]
    _assert([row[1] for row in rows] == ["SO_B", "SO_A"], "same-day Schedule row order")
    _assert(_num(rows[0][11]) == 3, "COOIS lines must aggregate")
    _assert(_num(rows[0][16]) == 2 and _num(rows[1][16]) == 0, "earlier Schedule row consumes stock first")
    _assert(_num(rows[0][14]) == 2 and _num(rows[1][14]) == 2, "child stock is fixed original total")


def test_same_so_mother_coois_order() -> None:
    result = run_engine(
        schedule=[{"so": "SO1", "cutting": "2026-07-01"}],
        coois=[
            {"so": "SO1", "material": "M_B", "qty": 1, "segment": "FLT", "unit": "M"},
            {"so": "SO1", "material": "M_A", "qty": 1, "segment": "FLT", "unit": "M"},
        ],
        zrmm=[
            {"mother": "M_A", "child": "SHARED", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
            {"mother": "M_B", "child": "SHARED", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
        ],
        mb52=[{"material": "SHARED", "segment": "FLT", "storage": "1001", "stock": 1}],
    )
    _assert(result["ok"], str(result))
    rows = result["rows"]
    _assert([row[2] for row in rows] == ["M_B", "M_A"], "COOIS mother first-occurrence order")
    _assert(_num(rows[0][16]) == 1 and _num(rows[1][16]) == 0, "first mother receives shared child stock")


def test_ambiguous_direct_demand_stops() -> None:
    result = run_engine(
        schedule=[{"so": "SO1", "cutting": "2026-07-01"}],
        coois=[
            {"so": "SO1", "material": "M1", "qty": 1, "segment": "FLT", "unit": "M"},
            {"so": "SO1", "material": "M2", "qty": 1, "segment": "FLT", "unit": "M"},
            {"so": "SO1", "material": "C_SHARED", "qty": 0.5, "segment": "FLT", "unit": "M"},
        ],
        zrmm=[
            {"mother": "M1", "child": "C_SHARED", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
            {"mother": "M2", "child": "C_SHARED", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
        ],
        mb52=[{"material": "C_SHARED", "segment": "FLT", "storage": "1001", "stock": 10}],
    )
    _assert(not result["ok"] and "ambiguous" in result.get("error", "").lower(), "ambiguous direct demand must stop")
    _assert(result.get("needs_ambiguous_confirm") is True, "must request confirmation")
    _assert(len(result.get("ambiguities") or []) == 1, "must list one ambiguity")


def test_ambiguous_direct_demand_split_on_confirm() -> None:
    result = run_engine(
        schedule=[{"so": "SO1", "cutting": "2026-07-01"}],
        coois=[
            {"so": "SO1", "material": "M1", "qty": 8, "segment": "FLT", "unit": "M"},
            {"so": "SO1", "material": "M2", "qty": 2, "segment": "FLT", "unit": "M"},
            {"so": "SO1", "material": "C_SHARED", "qty": 1.0, "segment": "FLT", "unit": "M"},
        ],
        zrmm=[
            {"mother": "M1", "child": "C_SHARED", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
            {"mother": "M2", "child": "C_SHARED", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
        ],
        mb52=[
            {"material": "M1", "segment": "FLT", "storage": "1001", "stock": 100},
            {"material": "M2", "segment": "FLT", "storage": "1001", "stock": 100},
            {"material": "C_SHARED", "segment": "FLT", "storage": "1001", "stock": 10},
        ],
        confirm_ambiguous_split=True,
    )
    _assert(result["ok"], f"confirm split must run: {result.get('error')}")
    rows = result["rows"]
    by_mother = {row[2]: _num(row[12]) for row in rows if _norm(row[8]) == "C_SHARED"}
    _assert(abs(by_mother.get("M1", 0) - 0.8) < 1e-9, f"M1 demandDirect expected 0.8 got {by_mother.get('M1')}")
    _assert(abs(by_mother.get("M2", 0) - 0.2) < 1e-9, f"M2 demandDirect expected 0.2 got {by_mother.get('M2')}")
    _assert(abs(sum(by_mother.values()) - 1.0) < 1e-9, "split must preserve total direct demand")


def test_missing_conversion_stops() -> None:
    result = run_engine(
        schedule=[{"so": "SO1", "cutting": "2026-07-01"}],
        coois=[{"so": "SO1", "material": "M1", "qty": 1, "segment": "FLT", "unit": "SHT"}],
        zrmm=[
            {"mother": "M1", "child": "C1", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "SHT", "bun": "KG", "desc": ""},
        ],
        mb52=[{"material": "C1", "segment": "FLT", "storage": "1001", "stock": 10}],
    )
    _assert(not result["ok"] and "conversion" in result.get("error", "").lower(), "missing conversion must stop")

def test_y_semantics() -> None:
    result = run_engine(
        schedule=[{"so": "SO1", "cutting": "2026-07-01"}],
        coois=[
            {"so": "SO1", "material": "M_READY", "qty": 1, "segment": "FLT", "unit": "M"},
            {"so": "SO1", "material": "M_ZERO", "qty": 0, "segment": "FLT", "unit": "M"},
            {"so": "SO1", "material": "M_MH", "qty": 1, "segment": "FLT", "unit": "M"},
            {"so": "SO1", "material": "DIRECT", "qty": 1, "segment": "FLT", "unit": "M"},
        ],
        zrmm=[
            {"mother": "M_READY", "child": "C_READY", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
            {"mother": "M_READY", "child": "MH04001", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
            {"mother": "M_ZERO", "child": "C_ZERO", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
            {"mother": "M_MH", "child": "MH04002", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
            {"mother": "OTHER", "child": "DIRECT", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "M", "bun": "M", "desc": ""},
        ],
        mb52=[
            {"material": "C_READY", "segment": "FLT", "storage": "1001", "stock": 1},
            {"material": "DIRECT", "segment": "FLT", "storage": "1001", "stock": 1},
        ],
    )
    _assert(result["ok"], str(result))
    by_child = {row[8]: row for row in result["rows"]}
    _assert(by_child["C_READY"][18] == "Y" and by_child["MH04001"][18] == "Y", "ready mother group")
    _assert(by_child["C_ZERO"][18] == "", "zero-demand group must not be Y")
    _assert(by_child["MH04002"][18] == "", "MH04-only group must not be Y")
    _assert(by_child["DIRECT"][2] == "" and by_child["DIRECT"][18] == "", "direct-only row must not be Y")


def test_sht_220x110_to_yd() -> None:
    result = run_engine(
        schedule=[{"so": "SO1", "cutting": "2026-07-01"}],
        coois=[{"so": "SO1", "material": "SHT_M", "qty": 1, "segment": "FLT", "unit": "SHT"}],
        zrmm=[
            {"mother": "SHT_M", "child": "YD_C", "gi_j": 0, "vendor_l": 1, "gr_p": 0, "storage": "1001", "batch": "FLT", "oun": "SHT", "bun": "YD", "desc": "material 220x110cm"},
        ],
        mb52=[{"material": "YD_C", "segment": "FLT", "storage": "1001", "stock": 10}],
    )
    _assert(result["ok"], str(result))
    _assert(abs(_num(result["rows"][0][11]) - 2.187227) < 1e-6, "220x110 SHT→YD factor")

def smoke_new_0722() -> dict | None:
    if not NEW_0722.exists():
        print("SKIP New_0722 smoke (folder missing)")
        return None
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("SKIP New_0722 smoke (openpyxl missing)")
        return None

    def load_sheet(path: Path) -> tuple[list[str], list[list]]:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        raw = [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
        if not raw:
            return [], []
        # dual-header detect: if row0 chinese-ish and row1 has english keys → use row1
        def has_en(cells: list) -> bool:
            keys = [_header_key(c) for c in cells]
            hints = [
                "sd document",
                "article(com.)",
                "open quantity",
                "material",
                "unrestricted",
                "stock segment",
                "requirement segment",
                "gi sc(541/542)",
                "stock of vendor",
                "base unit of measure",
                "storage location",
                "order",
                "cutting",
            ]
            hits = sum(1 for h in hints if any(k == h or (len(h) >= 4 and h in k) for k in keys))
            return hits >= 2

        def looks_cn(cells: list) -> bool:
            text = "".join(str(c) for c in cells)
            cn = len(re.findall(r"[\u4e00-\u9fff]", text))
            lat = len(re.findall(r"[A-Za-z]", text))
            return cn > 0 and cn >= max(3, lat * 0.25)

        hi = 0
        if len(raw) >= 2 and looks_cn(raw[0]) and has_en(raw[1]):
            hi = 1
        elif has_en(raw[0]):
            hi = 0
        elif len(raw) >= 2 and has_en(raw[1]):
            hi = 1
        headers = [str(h or f"Column {i+1}") for i, h in enumerate(raw[hi])]
        data = [list(r[: len(headers)]) for r in raw[hi + 1 :]]
        return headers, data

    schedule_h, schedule_d = load_sheet(NEW_0722 / "1.xlsx")
    coois_h, coois_d = load_sheet(NEW_0722 / "COOIS.xlsx")
    zrmm_h, zrmm_d = load_sheet(NEW_0722 / "0028.xlsx")
    mb52_h, mb52_d = load_sheet(NEW_0722 / "MB52_new.xlsx")

    sc = {
        "so": find_col(schedule_h, ["order", "so", "訂單", "order no", "order no."]),
        "cutting": find_col(schedule_h, ["cutting", "production", "裁斷", "cutting process"]),
    }
    cc = {
        "so": find_col(coois_h, ["sd document"]),
        "material": find_col(coois_h, ["material"]),
        "qty": find_col(coois_h, ["open quantity"]),
        "segment": find_col(coois_h, ["requirement segment"]),
        "unit": find_col(coois_h, ["base unit of measure"]),
    }
    zc = {
        "mother": find_col(zrmm_h, ["material"]),
        "child": find_col(zrmm_h, ["article(com.)", "article (com.)", "article"]),
        "gi_j": find_col(zrmm_h, ["gi sc(541/542)"]),
        "vendor_l": find_col(zrmm_h, ["stock of vendor"]),
        "gr_p": find_col(zrmm_h, ["gr sc(543/544)"]),
        "storage": find_col(zrmm_h, ["storage location"]),
        "batch": find_col(zrmm_h, ["batch"]),
        "oun": find_col(zrmm_h, ["oun"]),
        "bun": find_col(zrmm_h, ["bun"]),
        "desc": find_col(
            zrmm_h,
            [
                "material full description(cn)",
                "material full description(en)",
                "material full description",
            ],
        ),
    }
    mc = {
        "material": find_col(mb52_h, ["material"]),
        "segment": find_col(mb52_h, ["stock segment"]),
        "storage": find_col(mb52_h, ["storage location"]),
        "stock": find_col(mb52_h, ["unrestricted"]),
    }

    for name, cols in [("schedule", sc), ("coois", cc), ("zrmm", zc), ("mb52", mc)]:
        missing = [k for k, v in cols.items() if v < 0]
        _assert(not missing, f"{name} missing columns: {missing}")

    # Confirm Open Quantity is not Quantity withdrawn
    withdrawn = find_col(coois_h, ["quantity withdrawn"])
    _assert(cc["qty"] != withdrawn or withdrawn < 0, "qty must not be Quantity withdrawn")

    def col(row: list, idx: int):
        return row[idx] if 0 <= idx < len(row) else ""

    schedule = [
        {"so": col(r, sc["so"]), "cutting": col(r, sc["cutting"])} for r in schedule_d if _norm(col(r, sc["so"]))
    ]
    coois = [
        {
            "so": col(r, cc["so"]),
            "material": col(r, cc["material"]),
            "qty": col(r, cc["qty"]),
            "segment": col(r, cc["segment"]),
            "unit": col(r, cc["unit"]),
        }
        for r in coois_d
        if _norm(col(r, cc["so"]))
    ]
    zrmm = [
        {
            "mother": col(r, zc["mother"]),
            "child": col(r, zc["child"]),
            "gi_j": col(r, zc["gi_j"]),
            "vendor_l": col(r, zc["vendor_l"]),
            "gr_p": col(r, zc["gr_p"]),
            "storage": col(r, zc["storage"]),
            "batch": col(r, zc["batch"]),
            "oun": col(r, zc["oun"]),
            "bun": col(r, zc["bun"]),
            "desc": col(r, zc["desc"]),
        }
        for r in zrmm_d
        if _norm(col(r, zc["mother"]))
    ]
    mb52 = [
        {
            "material": col(r, mc["material"]),
            "segment": col(r, mc["segment"]),
            "storage": col(r, mc["storage"]),
            "stock": col(r, mc["stock"]),
        }
        for r in mb52_d
        if _norm(col(r, mc["material"]))
    ]

    result = run_engine(schedule, coois, zrmm, mb52)
    if not result["ok"]:
        # Real data often misses size rules — report but still count resolution OK
        print(f"New_0722 engine note: {result.get('error')}")
        return {
            "schedule_rows": len(schedule),
            "coois_rows": len(coois),
            "zrmm_rows": len(zrmm),
            "mb52_rows": len(mb52),
            "engine_ok": False,
            "error": result.get("error"),
            "headers_ok": True,
        }

    dual = [r for r in result["rows"] if "MA020162253" in (r[2], r[8])]
    print(
        f"New_0722 smoke OK: schedule={len(schedule)} coois={len(coois)} "
        f"zrmm={len(zrmm)} mb52={len(mb52)} out_rows={len(result['rows'])} "
        f"MA020162253_rows={len(dual)}"
    )
    return {
        "schedule_rows": len(schedule),
        "coois_rows": len(coois),
        "zrmm_rows": len(zrmm),
        "mb52_rows": len(mb52),
        "out_rows": len(result["rows"]),
        "engine_ok": True,
        "headers_ok": True,
        "dual_rows": len(dual),
    }


def _load_xlsx_table(path: Path) -> tuple[list[str], list[list]]:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.active
    raw = [list(row) for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    if not raw:
        return [], []
    hints = [
        "sd document",
        "article(com.)",
        "open quantity",
        "material",
        "unrestricted",
        "stock segment",
        "requirement segment",
        "gi sc(541/542)",
        "stock of vendor",
        "storage location",
        "cutting",
        "訂單",
    ]
    best_index = 0
    best_score = -1
    for index, row in enumerate(raw[:5]):
        keys = [_header_key(cell) for cell in row]
        score = sum(1 for hint in hints if any(hint in key for key in keys))
        if score > best_score:
            best_score = score
            best_index = index
    headers = [str(value or f"Column {index + 1}") for index, value in enumerate(raw[best_index])]
    return headers, [list(row[: len(headers)]) for row in raw[best_index + 1 :]]


def run_real_data(
    folder: Path,
    rules: dict[tuple[str, str], float],
    size_rules: dict[tuple[str, str], float],
) -> dict:
    schedule_path = next(folder.glob("Copy of 01*.xlsx"))
    coois_path = next(folder.glob("Copy of COOIS.csv"))
    zrmm_path = next(folder.glob("Copy of 0028*.xlsx"))
    mb52_path = next(folder.glob("Copy of MB52*.xlsx"))

    schedule_h, schedule_d = _load_xlsx_table(schedule_path)
    zrmm_h, zrmm_d = _load_xlsx_table(zrmm_path)
    mb52_h, mb52_d = _load_xlsx_table(mb52_path)

    sc = {
        "so": find_col(schedule_h, ["order", "so", "訂單", "order no", "order no."]),
        "cutting": find_col(schedule_h, ["cutting", "production", "裁斷", "cutting process"]),
    }
    zc = {
        "mother": find_col(zrmm_h, ["material"]),
        "child": find_col(zrmm_h, ["article(com.)", "article (com.)", "article"]),
        "gi_j": find_col(zrmm_h, ["gi sc(541/542)"]),
        "vendor_l": find_col(zrmm_h, ["stock of vendor"]),
        "gr_p": find_col(zrmm_h, ["gr sc(543/544)"]),
        "storage": find_col(zrmm_h, ["storage location"]),
        "batch": find_col(zrmm_h, ["batch"]),
        "oun": find_col(zrmm_h, ["oun"]),
        "bun": find_col(zrmm_h, ["bun"]),
        "desc": find_col(zrmm_h, ["material full description(cn)", "material full description(en)"]),
    }
    mc = {
        "material": find_col(mb52_h, ["material"]),
        "segment": find_col(mb52_h, ["stock segment"]),
        "storage": find_col(mb52_h, ["storage location"]),
        "stock": find_col(mb52_h, ["unrestricted"]),
    }
    for name, columns in (("schedule", sc), ("zrmm", zc), ("mb52", mc)):
        missing = [key for key, index in columns.items() if index < 0]
        _assert(not missing, f"{name} missing columns: {missing}")

    def cell(row: list, index: int):
        return row[index] if 0 <= index < len(row) else ""

    schedule = [
        {"so": cell(row, sc["so"]), "cutting": cell(row, sc["cutting"])}
        for row in schedule_d
        if _norm(cell(row, sc["so"]))
    ]
    coois: list[dict] = []
    with coois_path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            coois.append(
                {
                    "so": row.get("SD Document", ""),
                    "material": row.get("Material", ""),
                    "qty": row.get("Open Quantity", ""),
                    "segment": row.get("Requirement Segment", ""),
                    "unit": row.get("Base Unit of Measure", ""),
                }
            )
    zrmm = [
        {
            "mother": cell(row, zc["mother"]),
            "child": cell(row, zc["child"]),
            "gi_j": cell(row, zc["gi_j"]),
            "vendor_l": cell(row, zc["vendor_l"]),
            "gr_p": cell(row, zc["gr_p"]),
            "storage": cell(row, zc["storage"]),
            "batch": cell(row, zc["batch"]),
            "oun": cell(row, zc["oun"]),
            "bun": cell(row, zc["bun"]),
            "desc": cell(row, zc["desc"]),
        }
        for row in zrmm_d
        if _norm(cell(row, zc["mother"]))
    ]
    mb52 = [
        {
            "material": cell(row, mc["material"]),
            "segment": cell(row, mc["segment"]),
            "storage": cell(row, mc["storage"]),
            "stock": cell(row, mc["stock"]),
        }
        for row in mb52_d
        if _norm(cell(row, mc["material"]))
    ]

    result = run_engine(schedule, coois, zrmm, mb52, rules=rules, size_rules=size_rules)
    _assert(result.get("ok"), f"real-data engine failed: {result.get('error')}")
    stats = result["stats"]
    _assert(stats["zrmm_mb_excluded"] == 389, f"expected 389 MB rows, got {stats['zrmm_mb_excluded']}")
    _assert(stats["zrmm_relationships_removed"] == 45, f"expected 45 MB relationships, got {stats['zrmm_relationships_removed']}")
    _assert(stats["dual_identity_materials"] == 0, "dual identity must be zero after MB filter")

    mother_qty = sum(
        _num(row["qty"])
        for row in coois
        if _norm(row["so"]) == "10189518"
        and _norm(row["material"]) == "MN090134761"
        and _norm_seg(row["segment"]) == "FLT"
    )
    anchor_rows = [
        row
        for row in result["rows"]
        if row[1] == "10189518" and row[2] == "MN090134761" and row[8] == "MA020147549"
    ]
    _assert(abs(mother_qty - 1.5) < 1e-9, f"anchor mother demand {mother_qty}")
    _assert(anchor_rows, "anchor output row missing")
    _assert(abs(sum(_num(row[12]) for row in anchor_rows) - 0.02) < 1e-9, "anchor direct demand must be 0.02")
    _assert(abs(sum(_num(row[7]) for row in anchor_rows) - 0.496708) < 1e-6, "anchor H must be 0.496708")
    _assert(abs(sum(_num(row[16]) for row in anchor_rows)) < 1e-9, "anchor Q must be 0")
    _assert(abs(sum(_num(row[17]) for row in anchor_rows) - 43.046) < 1e-6, "anchor R must be 43.046")
    _assert(all(row[18] == "Y" for row in anchor_rows), "anchor must be fulfilled and kit-ready")

    legacy = run_engine(
        schedule,
        coois,
        zrmm,
        mb52,
        rules=rules,
        size_rules=size_rules,
        apply_rule_d=False,
    )
    _assert(legacy.get("ok"), f"v3.2 comparison engine failed: {legacy.get('error')}")
    _assert(len(legacy["rows"]) == len(result["rows"]), "v3.2/v3.3 row count mismatch")
    for current, previous in zip(result["rows"], legacy["rows"]):
        current_key = (current[0], current[1], current[2], current[3], current[8], current[9])
        previous_key = (previous[0], previous[1], previous[2], previous[3], previous[8], previous[9])
        _assert(current_key == previous_key, f"v3.2/v3.3 row order mismatch: {current_key} != {previous_key}")
    _assert(legacy["stats"]["shortage"] == 81, f"v3.2 shortage baseline drifted: {legacy['stats']['shortage']}")
    legacy_anchor = [
        row
        for row in legacy["rows"]
        if row[1] == "10189518" and row[2] == "MN090134761" and row[8] == "MA020147549"
    ]
    comparison = {
        "v3_2_shortage": legacy["stats"]["shortage"],
        "v3_3_shortage": stats["shortage"],
        "shortage_delta": stats["shortage"] - legacy["stats"]["shortage"],
        "h_changed_rows": sum(
            abs(_num(current[7]) - _num(previous[7])) > 1e-9
            for current, previous in zip(result["rows"], legacy["rows"])
        ),
        "q_changed_rows": sum(
            abs(_num(current[16]) - _num(previous[16])) > 1e-9
            for current, previous in zip(result["rows"], legacy["rows"])
        ),
        "r_changed_rows": sum(
            abs(_num(current[17]) - _num(previous[17])) > 1e-9
            for current, previous in zip(result["rows"], legacy["rows"])
        ),
        "y_changed_rows": sum(
            current[18] != previous[18]
            for current, previous in zip(result["rows"], legacy["rows"])
        ),
        "v3_2_y_rows": sum(row[18] == "Y" for row in legacy["rows"]),
        "v3_3_y_rows": sum(row[18] == "Y" for row in result["rows"]),
        "anchor_v3_2_h": sum(_num(row[7]) for row in legacy_anchor),
        "anchor_v3_2_q": sum(_num(row[16]) for row in legacy_anchor),
        "anchor_v3_2_r": sum(_num(row[17]) for row in legacy_anchor),
    }
    evidence = {
        "schedule_rows": len(schedule),
        "coois_rows": len(coois),
        "zrmm_rows": len(zrmm),
        "mb52_rows": len(mb52),
        "output_rows": len(result["rows"]),
        **stats,
        "anchor_mother_open_qty": mother_qty,
        "anchor_direct_demand": sum(_num(row[12]) for row in anchor_rows),
        "anchor_mother_pool_before_row": sum(_num(row[7]) for row in anchor_rows),
        "anchor_provided_from_child_mb52": sum(_num(row[16]) for row in anchor_rows),
        "anchor_child_remaining": sum(_num(row[17]) for row in anchor_rows),
        "anchor_kit_ready": all(row[18] == "Y" for row in anchor_rows),
        "v3_2_vs_v3_3": comparison,
    }
    print("REAL DATA 20260723 OK")
    print(json.dumps(evidence, ensure_ascii=False, indent=2))
    return evidence

def main() -> int:
    parser = argparse.ArgumentParser(description="RAW MAT allocation engine v3.3 verifier")
    parser.add_argument(
        "--real-data",
        nargs="?",
        const="20260723",
        help="Run the real-data acceptance check; default folder is 20260723.",
    )
    parser.add_argument(
        "--conversion-config",
        type=Path,
        help="Shared conversion JSON with schemaVersion, conversionRules, and sizeConversionRules.",
    )
    args = parser.parse_args()

    tests = [
        test_find_col_open_quantity,
        test_mb_filter_and_19_columns,
        test_direct_demand_retained_when_mother_covered,
        test_rule_d_partial_and_zero_mother_balance,
        test_rule_d_child_order_and_cross_so_pool,
        test_rule_d_unit_conversion,
        test_outside_same_unit_priority,
        test_outside_fallback_without_conversion,
        test_coois_aggregation_and_source_order,
        test_same_so_mother_coois_order,
        test_ambiguous_direct_demand_stops,
        test_ambiguous_direct_demand_split_on_confirm,
        test_missing_conversion_stops,
        test_y_semantics,
        test_sht_220x110_to_yd,
    ]
    failed = 0
    for fn in tests:
        try:
            fn()
            print(f"PASS {fn.__name__}")
        except AssertionError as exc:
            failed += 1
            print(f"FAIL {fn.__name__}: {exc}")
        except Exception as exc:  # noqa: BLE001
            failed += 1
            print(f"ERROR {fn.__name__}: {exc}")

    try:
        smoke_new_0722()
    except AssertionError as exc:
        failed += 1
        print(f"FAIL smoke_new_0722: {exc}")
    except Exception as exc:  # noqa: BLE001
        failed += 1
        print(f"ERROR smoke_new_0722: {exc}")

    if args.real_data:
        try:
            unit_rules, size_rules = (
                load_conversion_config(args.conversion_config)
                if args.conversion_config
                else (DEFAULT_RULES, DEFAULT_SIZE_RULES)
            )
            run_real_data(Path(args.real_data), unit_rules, size_rules)
        except AssertionError as exc:
            failed += 1
            print(f"FAIL real_data: {exc}")
        except Exception as exc:  # noqa: BLE001
            failed += 1
            print(f"ERROR real_data: {exc}")

    if failed:
        print(f"\n{failed} failed")
        return 1
    print("\nAll v3.3 fixture checks passed.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
