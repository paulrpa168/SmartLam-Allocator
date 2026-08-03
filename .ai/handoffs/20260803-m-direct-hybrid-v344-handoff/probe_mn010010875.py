from __future__ import annotations

import importlib.util
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(r"D:/0.AI-Agent-Workspace/03_projects/RAW MAT Project")
VERIFY = ROOT / ".ai/handoffs/20260722-raw-mat-allocation-v3/verify_allocation_v3.py"
MOTHER = "MN010010875"
DATASETS = ["0731", "0730"]

spec = importlib.util.spec_from_file_location("v", VERIFY)
mod = importlib.util.module_from_spec(spec)
assert spec and spec.loader
spec.loader.exec_module(mod)
m = mod


def load_xlsx_table(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    raw = [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()

    def has_en(cells):
        keys = [m._header_key(c) for c in cells]
        hints = [
            "sd document",
            "article(com.)",
            "open quantity",
            "material",
            "unrestricted",
            "stock segment",
            "requirement segment",
            "gi sc(541/542)",
            "stock of vendor",
            "base unit of measure",
            "storage location",
            "order",
            "cutting",
        ]
        return sum(1 for h in hints if any(k == h or (len(h) >= 4 and h in k) for k in keys)) >= 2

    def looks_cn(cells):
        text = "".join(str(c) for c in cells)
        cn = len(re.findall(r"[\u4e00-\u9fff]", text))
        lat = len(re.findall(r"[A-Za-z]", text))
        return cn > 0 and cn >= max(3, lat * 0.25)

    hi = 0
    if len(raw) >= 2 and looks_cn(raw[0]) and has_en(raw[1]):
        hi = 1
    elif has_en(raw[0]):
        hi = 0
    elif len(raw) >= 2 and has_en(raw[1]):
        hi = 1
    headers = [str(h or f"Column {i + 1}") for i, h in enumerate(raw[hi])]
    return headers, [list(r[: len(headers)]) for r in raw[hi + 1 :]]


def cell(row, idx):
    return row[idx] if 0 <= idx < len(row) else ""


def pick_files(base: Path):
    xs = list(base.glob("*.xlsx"))
    coois = next(p for p in xs if "coois" in p.name.lower())
    zrmm = next(p for p in xs if "0028" in p.name)
    mb52 = next(p for p in xs if "mb52" in p.name.lower())
    schedule = next(
        p
        for p in xs
        if p not in (coois, zrmm, mb52)
    )
    return schedule, coois, zrmm, mb52


def probe(folder: str) -> None:
    base = ROOT / folder
    print("=" * 72)
    print(f"DATASET {folder}  mother={MOTHER}")
    schedule_path, coois_path, zrmm_path, mb52_path = pick_files(base)
    print("files:", schedule_path.name, coois_path.name, zrmm_path.name, mb52_path.name)

    schedule_h, schedule_d = load_xlsx_table(schedule_path)
    coois_h, coois_d = load_xlsx_table(coois_path)
    zrmm_h, zrmm_d = load_xlsx_table(zrmm_path)
    mb52_h, mb52_d = load_xlsx_table(mb52_path)

    sc = {
        "so": m.find_col(schedule_h, ["order", "so", "訂單", "order no", "order no."]),
        "cutting": m.find_col(schedule_h, ["cutting", "production", "裁斷", "cutting process"]),
    }
    cc = {
        "so": m.find_col(coois_h, ["sd document"]),
        "material": m.find_col(coois_h, ["material"]),
        "qty": m.find_col(coois_h, ["open quantity"]),
        "segment": m.find_col(coois_h, ["requirement segment"]),
        "unit": m.find_col(coois_h, ["base unit of measure"]),
    }
    zc = {
        "mother": m.find_col(zrmm_h, ["material"]),
        "child": m.find_col(zrmm_h, ["article(com.)", "article (com.)", "article"]),
        "gi_j": m.find_col(zrmm_h, ["gi sc(541/542)"]),
        "vendor_l": m.find_col(zrmm_h, ["stock of vendor"]),
        "gr_p": m.find_col(zrmm_h, ["gr sc(543/544)"]),
        "storage": m.find_col(zrmm_h, ["storage location"]),
        "batch": m.find_col(zrmm_h, ["batch"]),
        "oun": m.find_col(zrmm_h, ["oun"]),
        "bun": m.find_col(zrmm_h, ["bun"]),
        "desc": m.find_col(
            zrmm_h, ["material full description(cn)", "material full description(en)"]
        ),
    }
    mc = {
        "material": m.find_col(mb52_h, ["material"]),
        "segment": m.find_col(mb52_h, ["stock segment"]),
        "storage": m.find_col(mb52_h, ["storage location"]),
        "stock": m.find_col(mb52_h, ["unrestricted"]),
    }

    # --- raw source facts ---
    mother_n = m._norm(MOTHER)
    mb_rows = [
        r
        for r in mb52_d
        if m._norm(cell(r, mc["material"])) == mother_n
    ]
    print(f"\nMB52 mother rows: {len(mb_rows)}")
    by_seg_storage: dict[tuple[str, str], float] = defaultdict(float)
    by_seg: dict[str, float] = defaultdict(float)
    for r in mb_rows:
        seg = m._norm_seg(cell(r, mc["segment"]))
        st = m._norm(cell(r, mc["storage"]))
        qty = m._num(cell(r, mc["stock"]))
        by_seg_storage[(seg, st)] += qty
        by_seg[seg] += qty
    for (seg, st), qty in sorted(by_seg_storage.items()):
        print(f"  MB52 seg={seg!r} storage={st!r} unrestricted={qty}")
    print("  MB52 by segment totals:", dict(by_seg))

    coois_m = [
        r
        for r in coois_d
        if m._norm(cell(r, cc["material"])) == mother_n
    ]
    print(f"\nCOOIS mother open rows: {len(coois_m)}")
    so_open: dict[str, float] = defaultdict(float)
    seg_open: dict[str, float] = defaultdict(float)
    for r in coois_m:
        so = m._norm(cell(r, cc["so"]))
        seg = m._norm_seg(cell(r, cc["segment"]))
        qty = m._num(cell(r, cc["qty"]))
        unit = m._norm(cell(r, cc["unit"]))
        so_open[so] += qty
        seg_open[f"{seg}|{unit}"] += qty
    print(f"  unique SO: {len(so_open)}  total open={sum(so_open.values())}")
    for k, v in sorted(seg_open.items()):
        print(f"  open seg/unit {k}: {v}")
    print("  top SO open:", sorted(so_open.items(), key=lambda x: -x[1])[:8])

    schedule_sos = {m._norm(cell(r, sc["so"])) for r in schedule_d if m._norm(cell(r, sc["so"]))}
    coois_sos = set(so_open)
    in_sched = sorted(coois_sos & schedule_sos)
    not_in_sched = sorted(coois_sos - schedule_sos)
    print(f"\nSchedule overlap: mother SO in schedule={len(in_sched)} / coois SO={len(coois_sos)}")
    if not_in_sched[:5]:
        print("  sample SO not in schedule:", not_in_sched[:5])

    zrmm_m = [r for r in zrmm_d if m._norm(cell(r, zc["mother"])) == mother_n]
    print(f"\nZRMM0028 BOM children for mother: {len(zrmm_m)}")
    children = []
    for r in zrmm_m:
        child = m._norm(cell(r, zc["child"]))
        if child.startswith("MB"):
            continue
        children.append(
            {
                "child": child,
                "batch": cell(r, zc["batch"]),
                "oun": m._norm(cell(r, zc["oun"])),
                "bun": m._norm(cell(r, zc["bun"])),
                "gi": m._num(cell(r, zc["gi_j"])),
                "vendor": m._num(cell(r, zc["vendor_l"])),
                "gr": m._num(cell(r, zc["gr_p"])),
                "storage": m._norm(cell(r, zc["storage"])),
            }
        )
    for c in children:
        print(
            f"  child={c['child']} oun={c['oun']} bun={c['bun']} "
            f"GI={c['gi']} vendor={c['vendor']} GR={c['gr']} stor={c['storage']}"
        )

    # child MB52
    for c in children:
        ch = c["child"]
        rows = [r for r in mb52_d if m._norm(cell(r, mc["material"])) == ch]
        tot = sum(m._num(cell(r, mc["stock"])) for r in rows)
        print(f"  child MB52 {ch}: rows={len(rows)} unrestricted_sum={tot}")

    # --- engine ---
    schedule = [
        {"so": cell(r, sc["so"]), "cutting": cell(r, sc["cutting"])}
        for r in schedule_d
        if m._norm(cell(r, sc["so"]))
    ]
    coois = [
        {
            "so": cell(r, cc["so"]),
            "material": cell(r, cc["material"]),
            "qty": cell(r, cc["qty"]),
            "segment": cell(r, cc["segment"]),
            "unit": cell(r, cc["unit"]),
        }
        for r in coois_d
        if m._norm(cell(r, cc["so"]))
    ]
    zrmm = [
        {
            "mother": cell(r, zc["mother"]),
            "child": cell(r, zc["child"]),
            "gi_j": cell(r, zc["gi_j"]),
            "vendor_l": cell(r, zc["vendor_l"]),
            "gr_p": cell(r, zc["gr_p"]),
            "storage": cell(r, zc["storage"]),
            "batch": cell(r, zc["batch"]),
            "oun": cell(r, zc["oun"]),
            "bun": cell(r, zc["bun"]),
            "desc": cell(r, zc["desc"]),
        }
        for r in zrmm_d
        if m._norm(cell(r, zc["mother"]))
    ]
    mb52 = [
        {
            "material": cell(r, mc["material"]),
            "segment": cell(r, mc["segment"]),
            "storage": cell(r, mc["storage"]),
            "stock": cell(r, mc["stock"]),
        }
        for r in mb52_d
        if m._norm(cell(r, mc["material"]))
    ]

    res = m.run_engine(schedule, coois, zrmm, mb52, confirm_ambiguous_split=True)
    rows = [r for r in res["rows"] if m._norm(r[2]) == mother_n]
    print(f"\nENGINE output rows for mother: {len(rows)}")
    if not rows:
        print("  *** NO OUTPUT ROWS — mother never entered allocation result ***")
        # diagnose: schedule/coois join
        print("  diagnose: mother present in COOIS?", len(coois_m) > 0)
        print("  diagnose: mother present in ZRMM (non-MB child)?", len(children) > 0)
        return

    # columns: A cutting, B so, C mother, D batch, E unit, F plant, G outside, H pool,
    # I child, J child batch, K unit, L demandFromMother, M direct, N total, O child stock,
    # P avail, Q provided, R remain, Y
    print(
        "SO | F | G | H | child | L | M | N | O | P | Q | R | Y"
    )
    f_vals = set()
    g_vals = set()
    q_sum = 0.0
    l_sum = 0.0
    m_sum = 0.0
    for r in rows:
        f_vals.add(r[5])
        g_vals.add(r[6])
        l_sum += m._num(r[11])
        m_sum += m._num(r[12])
        q_sum += m._num(r[16])
        print(
            f"{r[1]} | F={r[5]} G={r[6]} H={r[7]} | {r[8]} | "
            f"L={r[11]} M={r[12]} N={r[13]} | O={r[14]} P={r[15]} Q={r[16]} R={r[17]} | Y={r[18]!r}"
        )
    print(f"\nSUMMARY F shown={f_vals} G shown={g_vals}")
    print(f"  sum L={l_sum} sum M={m_sum} sum Q={q_sum}")
    print(f"  shortage_count in full run={res.get('shortage', res.get('stats', {}))}")
    if isinstance(res, dict):
        for k in ("shortage", "mothers", "output_rows", "stats", "warnings"):
            if k in res:
                print(f"  res[{k}]={res[k]}")


def main():
    for folder in DATASETS:
        if (ROOT / folder).exists():
            probe(folder)


if __name__ == "__main__":
    main()
