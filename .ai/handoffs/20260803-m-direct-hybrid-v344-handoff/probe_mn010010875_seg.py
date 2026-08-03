from __future__ import annotations

import importlib.util
import re
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(r"D:/0.AI-Agent-Workspace/03_projects/RAW MAT Project")
VERIFY = ROOT / ".ai/handoffs/20260722-raw-mat-allocation-v3/verify_allocation_v3.py"
MOTHER = "MN010010875"
BASE = ROOT / "0731"

spec = importlib.util.spec_from_file_location("v", VERIFY)
mod = importlib.util.module_from_spec(spec)
assert spec and spec.loader
spec.loader.exec_module(mod)
m = mod


def load_xlsx_table(path: Path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    raw = [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()

    def has_en(cells):
        keys = [m._header_key(c) for c in cells]
        hints = [
            "sd document",
            "article(com.)",
            "open quantity",
            "material",
            "unrestricted",
            "stock segment",
            "requirement segment",
            "gi sc(541/542)",
            "stock of vendor",
            "base unit of measure",
            "storage location",
            "order",
            "cutting",
        ]
        return sum(1 for h in hints if any(k == h or (len(h) >= 4 and h in k) for k in keys)) >= 2

    def looks_cn(cells):
        text = "".join(str(c) for c in cells)
        cn = len(re.findall(r"[\u4e00-\u9fff]", text))
        lat = len(re.findall(r"[A-Za-z]", text))
        return cn > 0 and cn >= max(3, lat * 0.25)

    hi = 0
    if len(raw) >= 2 and looks_cn(raw[0]) and has_en(raw[1]):
        hi = 1
    elif has_en(raw[0]):
        hi = 0
    elif len(raw) >= 2 and has_en(raw[1]):
        hi = 1
    headers = [str(h or f"Column {i + 1}") for i, h in enumerate(raw[hi])]
    return headers, [list(r[: len(headers)]) for r in raw[hi + 1 :]]


def cell(row, idx):
    return row[idx] if 0 <= idx < len(row) else ""


def main() -> None:
    xs = list(BASE.glob("*.xlsx"))
    coois_path = next(p for p in xs if "coois" in p.name.lower())
    zrmm_path = next(p for p in xs if "0028" in p.name)
    mb52_path = next(p for p in xs if "mb52" in p.name.lower())
    schedule_path = next(p for p in xs if p not in (coois_path, zrmm_path, mb52_path))
    schedule_h, schedule_d = load_xlsx_table(schedule_path)
    coois_h, coois_d = load_xlsx_table(coois_path)
    zrmm_h, zrmm_d = load_xlsx_table(zrmm_path)
    mb52_h, mb52_d = load_xlsx_table(mb52_path)
    sc = {
        "so": m.find_col(schedule_h, ["order", "so", "訂單", "order no", "order no."]),
        "cutting": m.find_col(schedule_h, ["cutting", "production", "裁斷", "cutting process"]),
    }
    cc = {
        "so": m.find_col(coois_h, ["sd document"]),
        "material": m.find_col(coois_h, ["material"]),
        "qty": m.find_col(coois_h, ["open quantity"]),
        "segment": m.find_col(coois_h, ["requirement segment"]),
        "unit": m.find_col(coois_h, ["base unit of measure"]),
    }
    zc = {
        "mother": m.find_col(zrmm_h, ["material"]),
        "child": m.find_col(zrmm_h, ["article(com.)", "article (com.)", "article"]),
        "gi_j": m.find_col(zrmm_h, ["gi sc(541/542)"]),
        "vendor_l": m.find_col(zrmm_h, ["stock of vendor"]),
        "gr_p": m.find_col(zrmm_h, ["gr sc(543/544)"]),
        "storage": m.find_col(zrmm_h, ["storage location"]),
        "batch": m.find_col(zrmm_h, ["batch"]),
        "oun": m.find_col(zrmm_h, ["oun"]),
        "bun": m.find_col(zrmm_h, ["bun"]),
        "desc": m.find_col(
            zrmm_h, ["material full description(cn)", "material full description(en)"]
        ),
    }
    mc = {
        "material": m.find_col(mb52_h, ["material"]),
        "segment": m.find_col(mb52_h, ["stock segment"]),
        "storage": m.find_col(mb52_h, ["storage location"]),
        "stock": m.find_col(mb52_h, ["unrestricted"]),
    }
    schedule = [
        {"so": cell(r, sc["so"]), "cutting": cell(r, sc["cutting"])}
        for r in schedule_d
        if m._norm(cell(r, sc["so"]))
    ]
    coois = [
        {
            "so": cell(r, cc["so"]),
            "material": cell(r, cc["material"]),
            "qty": cell(r, cc["qty"]),
            "segment": cell(r, cc["segment"]),
            "unit": cell(r, cc["unit"]),
        }
        for r in coois_d
        if m._norm(cell(r, cc["so"]))
    ]
    zrmm = [
        {
            "mother": cell(r, zc["mother"]),
            "child": cell(r, zc["child"]),
            "gi_j": cell(r, zc["gi_j"]),
            "vendor_l": cell(r, zc["vendor_l"]),
            "gr_p": cell(r, zc["gr_p"]),
            "storage": cell(r, zc["storage"]),
            "batch": cell(r, zc["batch"]),
            "oun": cell(r, zc["oun"]),
            "bun": cell(r, zc["bun"]),
            "desc": cell(r, zc["desc"]),
        }
        for r in zrmm_d
        if m._norm(cell(r, zc["mother"]))
    ]
    mb52 = [
        {
            "material": cell(r, mc["material"]),
            "segment": cell(r, mc["segment"]),
            "storage": cell(r, mc["storage"]),
            "stock": cell(r, mc["stock"]),
        }
        for r in mb52_d
        if m._norm(cell(r, mc["material"]))
    ]
    res = m.run_engine(schedule, coois, zrmm, mb52, confirm_ambiguous_split=True)
    rows = [r for r in res["rows"] if m._norm(r[2]) == MOTHER]

    flt = [r for r in rows if abs(m._num(r[5]) - 446.1) < 1e-6]
    mtf = [r for r in rows if abs(m._num(r[5]) - 446.1) >= 1e-6]
    print("0731 segment split by F")
    print(
        "FLT-like rows",
        len(flt),
        "sumL",
        sum(m._num(r[11]) for r in flt),
        "sumQ",
        sum(m._num(r[16]) for r in flt),
        "unique SO",
        len({r[1] for r in flt}),
    )
    print(
        "MTF-like rows",
        len(mtf),
        "sumL",
        sum(m._num(r[11]) for r in mtf),
        "sumQ",
        sum(m._num(r[16]) for r in mtf),
        "unique SO",
        len({r[1] for r in mtf}),
    )
    print("MTF SO H/L/Q (non-MH04 preferred):")
    seen: set[str] = set()
    for r in mtf:
        so = str(r[1])
        if so in seen:
            continue
        seen.add(so)
        kids = [x for x in mtf if x[1] == r[1]]
        x = next((k for k in kids if not str(k[8]).startswith("MH04")), kids[0])
        print(f"  SO={x[1]} F={x[5]} H={x[7]} L={x[11]} Q={x[16]} Y={x[18]!r} child={x[8]}")

    open_by: dict[tuple[str, str], float] = {}
    for r in coois_d:
        if m._norm(cell(r, cc["material"])) != MOTHER:
            continue
        so = m._norm(cell(r, cc["so"]))
        seg = m._norm_seg(cell(r, cc["segment"]))
        open_by[(so, seg)] = open_by.get((so, seg), 0.0) + m._num(cell(r, cc["qty"]))
    print("\nCOOIS open totals")
    print("FLT", sum(v for (_, seg), v in open_by.items() if seg == "FLT"))
    print("MTF", sum(v for (_, seg), v in open_by.items() if seg == "MTF"))
    print("top open:")
    for (so, seg), qty in sorted(open_by.items(), key=lambda x: -x[1])[:12]:
        print(f"  {so} {seg}: {qty}")

    # child Q only when L>0
    q_rows = [r for r in rows if m._num(r[16]) > 0]
    print("\nRows with Q>0:", len(q_rows))
    for r in q_rows:
        print(f"  SO={r[1]} F={r[5]} H={r[7]} child={r[8]} L={r[11]} Q={r[16]} Y={r[18]!r}")


if __name__ == "__main__":
    main()
