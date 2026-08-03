from __future__ import annotations

import importlib.util
import re
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[3]
VERIFY = ROOT / ".ai/handoffs/20260722-raw-mat-allocation-v3/verify_allocation_v3.py"
BASE = ROOT / "0730"
MOTHER = "MN090157252"
OUT = Path(__file__).resolve().parent / "MN090157252_formula_review_v3.4.4.xlsx"
OUT_ALIAS = Path(__file__).resolve().parent / "MN090157252_formula_review.xlsx"

spec = importlib.util.spec_from_file_location("v", VERIFY)
mod = importlib.util.module_from_spec(spec)
assert spec is not None and spec.loader is not None
spec.loader.exec_module(mod)
m = mod


def load_xlsx_table(path: Path) -> tuple[list[str], list[list]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    raw = [[("" if c is None else c) for c in row] for row in ws.iter_rows(values_only=True)]
    wb.close()

    def has_en(cells: list) -> bool:
        keys = [m._header_key(c) for c in cells]
        hints = [
            "sd document",
            "article(com.)",
            "open quantity",
            "material",
            "unrestricted",
            "stock segment",
            "requirement segment",
            "gi sc(541/542)",
            "stock of vendor",
            "base unit of measure",
            "storage location",
            "order",
            "cutting",
        ]
        return sum(1 for h in hints if any(k == h or (len(h) >= 4 and h in k) for k in keys)) >= 2

    def looks_cn(cells: list) -> bool:
        text = "".join(str(c) for c in cells)
        cn = len(re.findall(r"[\u4e00-\u9fff]", text))
        lat = len(re.findall(r"[A-Za-z]", text))
        return cn > 0 and cn >= max(3, lat * 0.25)

    hi = 0
    if len(raw) >= 2 and looks_cn(raw[0]) and has_en(raw[1]):
        hi = 1
    elif has_en(raw[0]):
        hi = 0
    elif len(raw) >= 2 and has_en(raw[1]):
        hi = 1
    headers = [str(h or f"Column {i + 1}") for i, h in enumerate(raw[hi])]
    return headers, [list(r[: len(headers)]) for r in raw[hi + 1 :]]


def cell(row: list, idx: int):
    return row[idx] if 0 <= idx < len(row) else ""


def main() -> None:
    schedule_path = [
        p2
        for p2 in BASE.glob("*.xlsx")
        if "0028" not in p2.name and "COOIS" not in p2.name.upper() and "MB52" not in p2.name.upper()
    ][0]
    coois_path = next(BASE.glob("COOIS*.xlsx"))
    zrmm_path = next(BASE.glob("0028*.xlsx"))
    mb52_path = next(BASE.glob("MB52*.xlsx"))
    schedule_h, schedule_d = load_xlsx_table(schedule_path)
    coois_h, coois_d = load_xlsx_table(coois_path)
    zrmm_h, zrmm_d = load_xlsx_table(zrmm_path)
    mb52_h, mb52_d = load_xlsx_table(mb52_path)
    sc = {
        "so": m.find_col(schedule_h, ["order", "so", "訂單", "order no", "order no."]),
        "cutting": m.find_col(schedule_h, ["cutting", "production", "裁斷", "cutting process"]),
    }
    cc = {
        "so": m.find_col(coois_h, ["sd document"]),
        "material": m.find_col(coois_h, ["material"]),
        "qty": m.find_col(coois_h, ["open quantity"]),
        "segment": m.find_col(coois_h, ["requirement segment"]),
        "unit": m.find_col(coois_h, ["base unit of measure"]),
    }
    zc = {
        "mother": m.find_col(zrmm_h, ["material"]),
        "child": m.find_col(zrmm_h, ["article(com.)", "article (com.)", "article"]),
        "gi_j": m.find_col(zrmm_h, ["gi sc(541/542)"]),
        "vendor_l": m.find_col(zrmm_h, ["stock of vendor"]),
        "gr_p": m.find_col(zrmm_h, ["gr sc(543/544)"]),
        "storage": m.find_col(zrmm_h, ["storage location"]),
        "batch": m.find_col(zrmm_h, ["batch"]),
        "oun": m.find_col(zrmm_h, ["oun"]),
        "bun": m.find_col(zrmm_h, ["bun"]),
        "desc": m.find_col(
            zrmm_h, ["material full description(cn)", "material full description(en)"]
        ),
    }
    mc = {
        "material": m.find_col(mb52_h, ["material"]),
        "segment": m.find_col(mb52_h, ["stock segment"]),
        "storage": m.find_col(mb52_h, ["storage location"]),
        "stock": m.find_col(mb52_h, ["unrestricted"]),
    }
    schedule = [
        {"so": cell(r, sc["so"]), "cutting": cell(r, sc["cutting"])}
        for r in schedule_d
        if m._norm(cell(r, sc["so"]))
    ]
    coois = [
        {
            "so": cell(r, cc["so"]),
            "material": cell(r, cc["material"]),
            "qty": cell(r, cc["qty"]),
            "segment": cell(r, cc["segment"]),
            "unit": cell(r, cc["unit"]),
        }
        for r in coois_d
        if m._norm(cell(r, cc["so"]))
    ]
    zrmm = [
        {
            "mother": cell(r, zc["mother"]),
            "child": cell(r, zc["child"]),
            "gi_j": cell(r, zc["gi_j"]),
            "vendor_l": cell(r, zc["vendor_l"]),
            "gr_p": cell(r, zc["gr_p"]),
            "storage": cell(r, zc["storage"]),
            "batch": cell(r, zc["batch"]),
            "oun": cell(r, zc["oun"]),
            "bun": cell(r, zc["bun"]),
            "desc": cell(r, zc["desc"]),
        }
        for r in zrmm_d
        if m._norm(cell(r, zc["mother"]))
    ]
    mb52 = [
        {
            "material": cell(r, mc["material"]),
            "segment": cell(r, mc["segment"]),
            "storage": cell(r, mc["storage"]),
            "stock": cell(r, mc["stock"]),
        }
        for r in mb52_d
        if m._norm(cell(r, mc["material"]))
    ]
    res = m.run_engine(schedule, coois, zrmm, mb52, confirm_ambiguous_split=True)
    rows = [r for r in res["rows"] if m._norm(r[2]) == MOTHER]

    letters = [get_column_letter(i) for i in range(1, 20)]
    zh = [
        "cutting（生產日）",
        "so（訂單）",
        "母材料",
        "母材料 BATCH",
        "母單位",
        "廠內母材料庫存",
        "廠外母材料庫存",
        "本列配發前母料可用池",
        "子材料",
        "子材料 BATCH",
        "子單位",
        "子材料需求",
        "子料直接需求",
        "需求合計",
        "子材料庫存",
        "可配發庫存",
        "能提供數量",
        "扣減後剩餘",
        "合貼備料齊套（Y）",
    ]
    formulas = [
        "排程 cutting",
        "銷售訂單",
        "C 母材料",
        "D 母材料 BATCH",
        "E 母單位",
        "F 廠內母材料庫存（固定）",
        "G 廠外 pick-once（固定）",
        "H：母池本列前；L=0 且 M>0 時 M 不扣此池",
        "I 子材料",
        "J 子材料 BATCH",
        "K 子單位",
        "L 子材料需求＝qtyExpand×換算",
        "M 子料直接需求（永遠顯示）",
        "N＝L+M",
        "O 子材料庫存（固定）",
        "P 可配發庫存（本列前）",
        "Q：L=0且M>0 → 0；L>0 → Rule D 未蓋完的 N",
        "R＝P−Q；L=0 時再扣 M（R=P−Q−M）",
        "Y：非MH04且L>0 皆 母蓋L+Q≥L；純M不標Y",
    ]

    thin = Border(
        left=Side(style="thin", color="B0B0B0"),
        right=Side(style="thin", color="B0B0B0"),
        top=Side(style="thin", color="B0B0B0"),
        bottom=Side(style="thin", color="B0B0B0"),
    )
    fill_letter = PatternFill("solid", fgColor="1F4E79")
    fill_name = PatternFill("solid", fgColor="2E75B6")
    fill_formula = PatternFill("solid", fgColor="FFF2CC")
    fill_ma = PatternFill("solid", fgColor="E2EFDA")
    font_white = Font(bold=True, color="FFFFFF", name="Microsoft JhengHei", size=10)
    font_f = Font(name="Microsoft JhengHei", size=9)
    font_d = Font(name="Microsoft JhengHei", size=10)
    wrap = Alignment(wrap_text=True, vertical="top")
    center = Alignment(wrap_text=True, vertical="center", horizontal="center")
    widths = [14, 12, 14, 12, 8, 14, 14, 16, 14, 12, 8, 12, 12, 12, 12, 12, 12, 12, 18]

    wb = Workbook()
    ws = wb.active
    ws.title = "定案公式_v344_MN090157252"
    ws["A1"] = "版本"
    ws["B1"] = "v3.4.4 混合規則（定案）"
    ws["A2"] = "母材料 C"
    ws["B2"] = MOTHER
    ws["A3"] = "黃列"
    ws["B3"] = "定案公式。綠列=MA020165546。資料=0730 引擎現行輸出。"
    ws.merge_cells("B3:S3")
    ws["A4"] = "混合規則"
    ws["B4"] = "L=0且M>0 → 母池不抵M、Q=0、R=P−M；L>0 → Rule D 蓋 N=L+M；Y 只看 L"
    ws.merge_cells("B4:S4")
    for col, letter in enumerate(letters, 1):
        c = ws.cell(6, col, letter)
        c.fill = fill_letter
        c.font = font_white
        c.alignment = center
        c.border = thin
    for col, name in enumerate(zh, 1):
        c = ws.cell(7, col, name)
        c.fill = fill_name
        c.font = font_white
        c.alignment = center
        c.border = thin
    for col, formula in enumerate(formulas, 1):
        c = ws.cell(8, col, formula)
        c.fill = fill_formula
        c.font = font_f
        c.alignment = wrap
        c.border = thin
    ws.row_dimensions[8].height = 90
    for r_i, row in enumerate(rows, 9):
        for c_i, val in enumerate(row, 1):
            cell_val = val
            if c_i in (6, 7, 8, 12, 13, 14, 15, 16, 17, 18) and val not in ("", None):
                try:
                    cell_val = float(str(val).replace(",", ""))
                except Exception:
                    cell_val = val
            c = ws.cell(r_i, c_i, cell_val)
            c.font = font_d
            c.border = thin
            if row[8] == "MA020165546":
                c.fill = fill_ma
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A9"

    ws3 = wb.create_sheet("欄位與名詞說明")
    legend = [
        ["代號", "名稱", "定案公式"],
        *[[letters[i], zh[i], formulas[i]] for i in range(19)],
        ["", "", ""],
        ["狀態", "內容", ""],
        ["已定", "混合規則 A（逐子料列）+ Y 只看 L + 頂部公式/匯出", ""],
    ]
    for r_i, row in enumerate(legend, 1):
        for c_i, val in enumerate(row, 1):
            c = ws3.cell(r_i, c_i, val)
            c.font = Font(bold=(r_i == 1), name="Microsoft JhengHei", size=10)
            c.border = thin
    ws3.column_dimensions["A"].width = 8
    ws3.column_dimensions["B"].width = 24
    ws3.column_dimensions["C"].width = 60

    wb.save(OUT)
    print(OUT, "rows", len(rows))
    try:
        wb.save(OUT_ALIAS)
        print("also wrote", OUT_ALIAS)
    except PermissionError:
        print("alias locked (open in Excel); kept", OUT.name)

    for r in rows:
        if r[8] == "MA020165546":
            print(r[1], "L", r[11], "M", r[12], "Q", r[16], "R", r[17], "Y", r[18])


if __name__ == "__main__":
    main()
